import os
import json
import asyncio
import aiohttp
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
from crawl4ai import AsyncWebCrawler
from crawl4ai.extraction_strategy import JsonCssExtractionStrategy
from crawl4ai.async_crawler_strategy import AsyncPlaywrightCrawlerStrategy
from bs4 import BeautifulSoup

class AmazonScraper:
    def __init__(self, output_dir="output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.images_dir = self.output_dir / "images"
        self.images_dir.mkdir(exist_ok=True)
        
        # Updated schema with precise selectors
        self.schema = {
            "name": "Amazon Product",
            "baseSelector": "body",
            "fields": [
                {
                    "name": "title",
                    "selector": "#productTitle",
                    "type": "text",
                    "transform": "strip"
                },
                {
                    "name": "price",
                    "selector": ".a-price .a-offscreen",
                    "type": "text",
                    "transform": "strip"
                },
                {
                    "name": "description",
                    "selector": "#feature-bullets .a-list-item",
                    "type": "text",
                    "transform": "strip"
                }
            ]
        }

    async def process_excel(self, input_file: str):
        """Process Excel file containing ASINs"""
        try:
            print(f"Opening Excel file: {input_file}")
            wb = load_workbook(filename=input_file)
            ws = wb.active
            
            # Find the ASIN column (should be column 5 based on your Excel)
            asin_column = 5  # This is column D in Excel
            
            # Extract ASINs from the ASIN column, starting from row 2
            asins = []
            for row in ws.iter_rows(min_row=2, min_col=asin_column, max_col=asin_column):
                cell_value = row[0].value
                if cell_value:
                    asin = str(cell_value).strip()
                    # Only include ASINs that start with B0
                    if asin and asin.startswith('B0'):
                        asins.append(asin)
                        print(f"Found valid ASIN: {asin}")
                    else:
                        print(f"Skipping invalid ASIN: {asin}")
            
            print(f"Found {len(asins)} valid ASINs to process")
            
            # Process each ASIN
            results = []
            async with AsyncWebCrawler(verbose=True) as crawler:
                for asin in asins:
                    url = f"https://www.amazon.com/dp/{asin}"
                    print(f"Processing ASIN: {asin} - URL: {url}")
                    
                    try:
                        session_id = f"amazon_session_{asin}"
                        
                        # Enhanced wait condition for image
                        wait_for = """() => {
                            const img = document.querySelector('#landingImage, #imgBlkFront');
                            if (!img) return false;
                            if (!img.complete) return false;
                            if (!img.naturalWidth) return false;
                            console.log('Image found:', img.src);
                            return true;
                        }"""
                        
                        result = await crawler.arun(
                            url=url,
                            session_id=session_id,
                            wait_for=wait_for,
                            extraction_strategy=JsonCssExtractionStrategy(self.schema),
                            bypass_cache=True,
                            headless=False,
                            timeout=60000
                        )
                        
                        if result.success:
                            # First get the structured data from Crawl4AI
                            data = json.loads(result.extracted_content) if result.extracted_content else [{}]
                            data_dict = data[0] if data else {}
                            print(f"Crawl4AI extracted data: {data_dict}")  # Debug print
                            
                            # Then use BeautifulSoup for image URL extraction
                            soup = BeautifulSoup(result.html, 'html.parser')
                            print("Parsing HTML with BeautifulSoup...")
                            
                            # Try to find the image element
                            img = soup.select_one('#landingImage, #imgBlkFront')
                            print(f"Found image element: {img}")
                            
                            image_url = None
                            if img:
                                # Try different attributes in order of preference
                                for attr in ['data-old-hires', 'data-a-dynamic-image', 'src']:
                                    if attr in img.attrs:
                                        if attr == 'data-a-dynamic-image':
                                            try:
                                                urls = json.loads(img[attr])
                                                image_url = list(urls.keys())[0]
                                                print(f"Found image URL from {attr}: {image_url}")
                                                break
                                            except:
                                                continue
                                        else:
                                            image_url = img[attr]
                                            print(f"Found image URL from {attr}: {image_url}")
                                            break
                            
                            print(f"Final image URL: {image_url}")
                            
                            # Combine the data
                            data_dict['image_url'] = image_url
                            
                            results.append({
                                "asin": asin,
                                "data": data_dict,
                                "error": None
                            })
                            
                            await crawler.crawler_strategy.kill_session(session_id)
                            
                    except Exception as e:
                        print(f"Error processing ASIN {asin}: {e}")
                        results.append({
                            "asin": asin,
                            "data": {},
                            "error": str(e)
                        })
            
            # Create output Excel file
            output_wb = Workbook()
            output_ws = output_wb.active
            
            # Write headers
            headers = ["ASIN", "Title", "Price", "Description", "Image URL", "Error"]
            for col, header in enumerate(headers, 1):
                output_ws.cell(row=1, column=col, value=header)
            
            # Write results
            for row, result in enumerate(results, 2):
                output_ws.cell(row=row, column=1, value=result["asin"])
                output_ws.cell(row=row, column=2, value=result["data"].get("title", ""))
                output_ws.cell(row=row, column=3, value=result["data"].get("price", ""))
                output_ws.cell(row=row, column=4, value=result["data"].get("description", ""))
                output_ws.cell(row=row, column=5, value=result["data"].get("image_url", ""))
                output_ws.cell(row=row, column=6, value=result["error"])
            
            # Save output file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = self.output_dir / f"amazon_data_{timestamp}.xlsx"
            output_wb.save(output_file)
            
            return output_file
            
        except Exception as e:
            print(f"Error processing Excel file: {e}")
            raise

    async def download_main_image(self, session, asin, image_data):
        """Download the main product image"""
        try:
            # Try to get high-res image URL from data-a-dynamic-image
            if 'data-a-dynamic-image' in image_data:
                # Parse the JSON string containing image URLs
                image_urls = json.loads(image_data['data-a-dynamic-image'])
                # Get the URL of the largest image (last in the list)
                image_url = list(image_urls.keys())[-1]
            else:
                # Fallback to regular src
                image_url = image_data['src']
            
            # Download the image
            async with session.get(image_url) as response:
                if response.status == 200:
                    filename = f"{asin}_main.jpg"
                    filepath = self.images_dir / filename
                    with open(filepath, 'wb') as f:
                        f.write(await response.read())
                    return str(filepath)
                
        except Exception as e:
            print(f"Error downloading image for ASIN {asin}: {e}")
        return None

async def main():
    try:
        input_file = "amzAMZ_OV_test_151.xlsx"
        
        scraper = AmazonScraper(output_dir="output")
        output_file = await scraper.process_excel(input_file)
        print(f"Processing complete. Results saved to: {output_file}")
        
    except Exception as e:
        print(f"Error in main execution: {e}")

if __name__ == "__main__":
    asyncio.run(main())